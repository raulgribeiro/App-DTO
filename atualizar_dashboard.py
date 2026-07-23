"""
Atualizar Dashboard DTO — script único
=======================================

O que faz, em ordem:
  1. Le a planilha exportada do Microsoft Forms (.xlsx)
  2. Extrai e limpa os dados das colunas B, F, J, L, M, O, P, R, Q:AK, AL, AM, AN
  3. Substitui os dados dentro do próprio index.html (a linha "const dtoRaw = [...]")
  4. Atualiza a data em "Atualizado em ..."
  5. Faz commit e push para o GitHub (o GitHub Pages publica sozinho)

USO:
  python atualizar_dashboard.py "caminho/para/export_do_forms.xlsx"

Se nenhum caminho for passado, o script usa o caminho padrão configurado em
DEFAULT_XLSX_PATH logo abaixo (já apontando para o arquivo da Clealco no
OneDrive). Se não achar nada lá, procura o .xlsx mais recente na pasta do script.

CONFIGURAÇÃO (ajuste na primeira vez que usar, ou pelo menos confira):
  - GIT_REMOTE_NAME: nome do remoto git (normalmente 'origin')
  - GIT_BRANCH: branch publicada pelo GitHub Pages (normalmente 'main')

Pré-requisito: rode este script de dentro da pasta do repositório git clonado de
https://github.com/raulgribeiro/App-DTO (remoto 'origin' apontando pra lá, branch 'main').
"""

import sys
import re
import json
import glob
import subprocess
from pathlib import Path
from datetime import datetime

import openpyxl

# ------------------ CONFIGURAÇÃO ------------------
REPO_DIR = Path(__file__).parent.resolve()
INDEX_FILE = REPO_DIR / "index.html"
GIT_REMOTE_NAME = "origin"
GIT_BRANCH = "main"
COMMIT_MSG_PREFIX = "Atualizacao automatica DTO"

# Caminho padrão do arquivo exportado do Forms (ajuste se mudar de lugar).
# Usado automaticamente quando o script é chamado sem argumento.
DEFAULT_XLSX_PATH = r"C:\Users\raulribeiro\OneDrive - CLEALCO AÇÚCAR E ÁLCOOL S.A\DTO Excel\Teste\Diagnóstico de Trabalho Operacional (DTO) SF 25_26.xlsx"
# ---------------------------------------------------

# Colunas com o nome da(s) IT(s) treinada(s), dependendo da área escolhida no Forms
IT_COLS = ['S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC',
           'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AK']


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace('\xa0', ' ').strip()
        return v if v else None
    return v


def encontrar_planilha(caminho_informado):
    if caminho_informado:
        p = Path(caminho_informado)
        if not p.exists():
            raise FileNotFoundError(f"Arquivo nao encontrado: {p}")
        return p

    # Tenta o caminho padrão configurado no topo do script
    if DEFAULT_XLSX_PATH:
        p_default = Path(DEFAULT_XLSX_PATH)
        if p_default.exists():
            return p_default

    candidatos = sorted(glob.glob(str(REPO_DIR / "*.xlsx")), key=lambda p: Path(p).stat().st_mtime, reverse=True)
    if not candidatos:
        raise FileNotFoundError(
            "Nenhum .xlsx encontrado (nem no caminho padrão, nem na pasta do script).\n"
            "Informe o caminho como argumento:\n"
            "  python atualizar_dashboard.py caminho/do/arquivo.xlsx"
        )
    return Path(candidatos[0])


def extrair_dados(caminho_xlsx):
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    ws = wb['Sheet1']
    registros = []
    for r in range(2, ws.max_row + 1):
        get = lambda c: clean(ws[f'{c}{r}'].value)
        data_raw = get('B')
        if data_raw is None:
            continue
        try:
            data_str = data_raw.strftime('%Y-%m-%d') if hasattr(data_raw, 'strftime') else str(data_raw)[:10]
        except Exception:
            data_str = str(data_raw)[:10]

        area_agricola = get('M')
        area_ind_cle = get('O')
        area_ind_qrz = get('P')
        area_manut = get('R')
        area_detalhe = area_agricola or area_ind_cle or area_ind_qrz or area_manut or None

        its = []
        for c in IT_COLS:
            v = get(c)
            if v:
                its.extend([x.strip() for x in v.split(';') if x.strip()])

        registros.append({
            'id': get('A'),
            'data': data_str,
            'avaliador': get('F'),
            'unidade': get('J'),
            'area': get('L'),
            'areaDetalhe': area_detalhe,
            'turno': get('AL'),
            'cargo': get('AM'),
            'colaborador': get('AN'),
            'its': its,
            'revisaoIT': get('BD'),
            'pontosRevisao': get('BE'),
        })
    return registros


def atualizar_index_html(registros):
    html = INDEX_FILE.read_text(encoding='utf-8')
    data_json = json.dumps(registros, ensure_ascii=False)

    novo_html, n = re.subn(
        r'const dtoRaw = \[.*?\];',
        f'const dtoRaw = {data_json};',
        html,
        count=1,
        flags=re.DOTALL,
    )
    if n == 0:
        raise RuntimeError("Nao encontrei 'const dtoRaw = [...]' no index.html — verifique o arquivo.")

    agora = datetime.now().strftime('%d/%m/%Y %H:%M')
    novo_html, n2 = re.subn(
        r"document\.getElementById\('lastUpdate'\)\.textContent = 'Atualizado em .*?';",
        f"document.getElementById('lastUpdate').textContent = 'Atualizado em {agora}';",
        novo_html,
        count=1,
    )

    INDEX_FILE.write_text(novo_html, encoding='utf-8')
    print(f"index.html atualizado: {len(registros)} registros, {agora}")


def rodar_git(cmd):
    print(f"$ git {' '.join(cmd)}")
    result = subprocess.run(["git"] + cmd, cwd=REPO_DIR, capture_output=True, text=True)
    if result.stdout.strip():
        print(result.stdout.strip())
    if result.returncode != 0:
        print(result.stderr.strip())
    return result


def main():
    caminho_arg = sys.argv[1] if len(sys.argv) > 1 else None
    planilha = encontrar_planilha(caminho_arg)
    print(f"Usando planilha: {planilha.name}")

    registros = extrair_dados(planilha)
    atualizar_index_html(registros)

    rodar_git(["add", "index.html"])
    msg = f"{COMMIT_MSG_PREFIX} - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    commit = rodar_git(["commit", "-m", msg])
    if "nothing to commit" in (commit.stdout + commit.stderr):
        print("Nada de novo para commitar (dados identicos aos ja publicados).")
        return
    rodar_git(["push", GIT_REMOTE_NAME, GIT_BRANCH])
    print("\nPronto! O GitHub Pages deve atualizar o site em alguns instantes.")


if __name__ == "__main__":
    main()
